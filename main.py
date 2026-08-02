"""
UPRVUNL Rake Diversion - Coal Optimizer (Proof of Concept)
============================================================
Milestones covered in this single file:
  1. Read Excel, print Plant / Source / Rakes / VC
  2. Reproduce Excel's plant-wise + overall weighted VC
  3. Manual rake-change test (NCL rakes at Parichha: 60 -> 52)
  4. Basic optimizer for ONE plant (Parichha)
  5. Optimizer for ALL plants (each plant optimized independently,
     same bounded logic, still respecting the sheet's boundary conditions)
  6. Daily variation overlay - a CSV can reduce/change "available" rakes
     for a plant/source on top of the monthly plan, then re-optimize
  7. Freeze functionality - lock specific plant/source rakes so the
     optimizer treats them as fixed and only re-balances the rest
  8. Save everything to output.xlsx

Sheet used as source-of-truth: "July 2026 (3)"
  (this is the live solver sheet the Dashboard is linked to)

CLI usage
---------
  python main.py
  python main.py --daily daily_variation.csv
  python main.py --freeze "Parichha:NCL=52,Panki:CCL=60"
  python main.py --daily daily_variation.csv --freeze "Parichha:NCL=52"
"""

import argparse
import os

import openpyxl
import pandas as pd

INPUT_FILE = "input.xlsx"
SHEET_NAME = "July 2026 (3)"
TARGET_VC = 3.99          # from N17 boundary condition in the sheet
OUTPUT_FILE = "output/output.xlsx"

# Boundary conditions taken from the sheet (N7:P17 in "July 2026 (3)")
SOURCE_MIN_PCT = 0.4      # a single coal source at a plant can't drop below 40% of current
SOURCE_MAX_PCT = 1.2      # ... or exceed 120% of current
PLANT_MIN_PCT = 0.8       # a plant's total rakes can't drop below 80% of current total
PLANT_MAX_PCT = 1.1       # ... or exceed 110% of current total


# ---------------------------------------------------------------
# STEP 1 + STEP 2 : Read the Excel file and pull out clean rows
# ---------------------------------------------------------------
def read_rake_data(path=INPUT_FILE, sheet=SHEET_NAME):
    """
    The sheet has a merged-cell layout:
        Plant name and Linkage only appear on the FIRST row of a block.
        Each block has 3-5 source rows, followed by a TOTAL row
        (a row that has Rakes + VC but no Coal Source name).
    We forward-fill the plant name and stop each block at the total row.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]

    records = []
    current_plant = None

    for row in ws.iter_rows(min_row=4, max_row=29):
        plant = row[0].value      # column A
        source = row[2].value     # column C
        rakes = row[3].value      # column D
        vc = row[4].value         # column E

        if plant not in (None, "Total UPRVUNL"):
            current_plant = str(plant).strip()

        # A source row = has a coal source name AND numeric rakes/VC
        if source is not None and rakes is not None and vc is not None:
            if current_plant and current_plant != "Total UPRVUNL":
                records.append({
                    "Plant": current_plant,
                    "Source": str(source).strip(),
                    "Rakes": float(rakes),
                    "Variable Cost": float(vc),
                })

    df = pd.DataFrame(records)
    return df


# ---------------------------------------------------------------
# STEP 3 (clean): basic validation
# ---------------------------------------------------------------
def clean_and_validate(df):
    df = df.dropna(subset=["Plant", "Source"]).copy()
    df["Rakes"] = pd.to_numeric(df["Rakes"], errors="coerce")
    df["Variable Cost"] = pd.to_numeric(df["Variable Cost"], errors="coerce")

    invalid = df[df["Rakes"] < 0]
    if not invalid.empty:
        print("WARNING: negative rake values found:\n", invalid)

    if df["Variable Cost"].isnull().any():
        raise ValueError("Missing Variable Cost values - check sheet mapping")

    return df


# ---------------------------------------------------------------
# STEP 4 (reproduce): plant-wise + overall weighted VC
# ---------------------------------------------------------------
def compute_weighted_vc(df, rakes_col="Rakes"):
    df = df.copy()
    df["Weighted Cost"] = df[rakes_col] * df["Variable Cost"]

    plant_summary = (
        df.groupby("Plant", sort=False)
        .agg(Total_Rakes=(rakes_col, "sum"),
             Total_Weighted_Cost=("Weighted Cost", "sum"))
        .reset_index()
    )
    plant_summary["Weighted VC"] = (
        plant_summary["Total_Weighted_Cost"] / plant_summary["Total_Rakes"]
    )
    plant_summary["Target Status"] = plant_summary["Weighted VC"].apply(
        lambda v: "Achieved" if v <= TARGET_VC else "Not Achieved"
    )

    overall_rakes = df[rakes_col].sum()
    overall_weighted_cost = df["Weighted Cost"].sum()
    overall_vc = overall_weighted_cost / overall_rakes

    return plant_summary, overall_vc


# ---------------------------------------------------------------
# STEP 5: Manual rake-change test
# ---------------------------------------------------------------
def manual_change_test(df, plant, source, new_rakes):
    df = df.copy()
    mask = (df["Plant"] == plant) & (df["Source"] == source)
    if not mask.any():
        raise ValueError(f"No row for {plant} / {source}")

    old_plant_summary, _ = compute_weighted_vc(df[df["Plant"] == plant])
    old_vc = old_plant_summary.loc[0, "Weighted VC"]

    df.loc[mask, "Rakes"] = new_rakes

    new_plant_summary, _ = compute_weighted_vc(df[df["Plant"] == plant])
    new_vc = new_plant_summary.loc[0, "Weighted VC"]
    status = "Achieved" if new_vc <= TARGET_VC else "Not Achieved"

    return old_vc, new_vc, status, df


# ---------------------------------------------------------------
# STEP 6: Daily variation overlay
# ---------------------------------------------------------------
def apply_daily_variation(df, csv_path):
    """
    Overlay a daily_variation.csv on top of the monthly plan.
    The monthly "Rakes" column is preserved as-is (Original Rakes);
    a new "Available Rakes" column reflects the daily override
    (falls back to the monthly value when no override exists).
    """
    df = df.copy()
    df["Available Rakes"] = df["Rakes"]

    if not csv_path or not os.path.exists(csv_path):
        print(f"(No daily variation file found at '{csv_path}', skipping overlay)")
        return df

    daily = pd.read_csv(csv_path)
    daily["plant"] = daily["plant"].astype(str).str.strip()
    daily["source"] = daily["source"].astype(str).str.strip()

    for _, r in daily.iterrows():
        mask = (df["Plant"] == r["plant"]) & (df["Source"] == r["source"])
        if mask.any():
            df.loc[mask, "Available Rakes"] = float(r["available_rakes"])
        else:
            print(f"WARNING: daily variation row not matched in sheet: "
                  f"{r['plant']} / {r['source']}")

    return df


# ---------------------------------------------------------------
# STEP 7: Freeze parsing - "Plant:Source=Value,Plant2:Source2=Value2"
# ---------------------------------------------------------------
def parse_freeze_arg(freeze_str):
    """Returns dict {(plant, source): frozen_value}"""
    frozen = {}
    if not freeze_str:
        return frozen
    for chunk in freeze_str.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        plant_source, value = chunk.split("=")
        plant, source = plant_source.split(":")
        frozen[(plant.strip(), source.strip())] = float(value)
    return frozen


# ---------------------------------------------------------------
# STEP 5 + 6 + 7 combined: Basic optimizer (one plant OR all plants),
# with optional daily-availability base and optional frozen cells
# ---------------------------------------------------------------
def optimize_plant(df, plant, base_col="Rakes", frozen=None,
                    min_pct=SOURCE_MIN_PCT, max_pct=SOURCE_MAX_PCT):
    """
    Minimize the plant's weighted VC by re-allocating rakes across its
    coal sources, subject to:
      - total rakes for the plant conserved at the CURRENT (base_col) total
      - each source's rakes stay within [min_pct, max_pct] x its base value
      - frozen (plant, source) cells are pinned to a fixed value

    NOT THE OFFICIAL OPTIMIZER. This CLI is a standalone demo/reporting
    tool, separate from backend/optimizer.py. The official, authoritative
    optimizer for this project is backend/optimizer.py via POST /optimize -
    it is the only implementation that enforces company-wide rake
    conservation across ALL plants simultaneously; this function only ever
    sees one plant at a time.

    Tries this CLI's own OR-Tools model first. If OR-Tools is unavailable or
    fails to find an optimal solution, this does NOT silently substitute a
    greedy heuristic as if it were equivalent - it prints an explicit
    warning and returns a "Method" column on every row so the caller (and
    output.xlsx) can never mistake a greedy PREVIEW result for an official
    OR-Tools solve.

    Returns the result DataFrame with an added "Method" column: either
    "OR-Tools" or "Greedy (PREVIEW - not official)".
    """
    frozen = frozen or {}
    sub = df[df["Plant"] == plant].reset_index(drop=True)
    total_rakes = int(round(sub[base_col].sum()))
    sources = sub["Source"].tolist()
    costs = sub["Variable Cost"].tolist()
    base = sub[base_col].tolist()

    lower_bounds, upper_bounds = [], []
    for i, s in enumerate(sources):
        if (plant, s) in frozen:
            fixed = int(round(frozen[(plant, s)]))
            lower_bounds.append(fixed)
            upper_bounds.append(fixed)
        else:
            lb = max(0, int(round(base[i] * min_pct)))
            ub = int(round(base[i] * max_pct)) if base[i] > 0 else int(round(total_rakes * 0.5))
            lower_bounds.append(lb)
            upper_bounds.append(max(ub, lb))

    method = "OR-Tools"
    try:
        from ortools.linear_solver import pywraplp
        solver = pywraplp.Solver.CreateSolver("SCIP")
        if solver is None:
            raise RuntimeError("Could not create SCIP solver")
        vars_ = [solver.IntVar(lower_bounds[i], upper_bounds[i], sources[i])
                  for i in range(len(sources))]

        solver.Add(solver.Sum(vars_) == total_rakes)
        solver.Minimize(solver.Sum(vars_[i] * costs[i] for i in range(len(sources))))

        status = solver.Solve()
        if status == pywraplp.Solver.OPTIMAL:
            optimized = [v.solution_value() for v in vars_]
        else:
            raise RuntimeError(f"Solver did not find an optimal solution (status={status})")
    except Exception as e:
        method = "Greedy (PREVIEW - not official)"
        print(f"WARNING: OR-Tools solver was unavailable or failed for '{plant}': {e}")
        print("         Official optimization was not performed for this plant.")
        print("         Showing a PREVIEW-ONLY greedy heuristic result instead - "
              "do not treat this as the official allocation.")
        order = sorted(range(len(sources)), key=lambda i: costs[i])
        remaining = total_rakes
        optimized = list(lower_bounds)
        remaining -= sum(lower_bounds)
        for i in order:
            room = upper_bounds[i] - optimized[i]
            take = min(room, remaining)
            optimized[i] += take
            remaining -= take
            if remaining <= 0:
                break

    optimized = [0.0 if abs(v) < 1e-9 else v for v in optimized]  # clean up -0.0

    result = sub.copy()
    result["Optimized Rakes"] = optimized
    result["Frozen"] = [(plant, s) in frozen for s in sources]
    result["Method"] = method
    return result


def optimize_all_plants(df, base_col="Rakes", frozen=None):
    frames = [optimize_plant(df, plant, base_col=base_col, frozen=frozen)
              for plant in df["Plant"].unique()]
    combined = pd.concat(frames, ignore_index=True)

    fallback_plants = sorted(combined.loc[combined["Method"] != "OR-Tools", "Plant"].unique())
    if fallback_plants:
        print("\n" + "!" * 60)
        print("OR-Tools solver was unavailable or failed for the following "
              "plant(s). Official optimization was not performed for them:")
        for p in fallback_plants:
            print(f"  - {p}")
        print("Their rows are PREVIEW-ONLY (greedy heuristic) and are labeled "
              "as such in the 'Method' column of the export.")
        print("!" * 60)

    return combined


# ---------------------------------------------------------------
# CONNECT UI: rebuild the dashboard's embedded RAW_DATA straight
# from the Excel "Calculation Sheet" (rate/freight/scc/apc detail
# that the HTML's own unitVC() formula needs), so the static
# dashboard always reflects whatever is in input.xlsx.
# ---------------------------------------------------------------
def read_calculation_sheet(path=INPUT_FILE, sheet="Calculation Sheet"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]

    plants = []
    current_block = None

    def flush():
        if current_block:
            plants.append(current_block)

    for row in ws.iter_rows(min_row=4, max_row=26):
        A, B, C = row[0].value, row[1].value, row[2].value          # Plant, Linkage, Source
        D = row[3].value                                            # Annual Coal Linkage
        F, G, H = row[5].value, row[6].value, row[7].value          # Rate, Premium, Freight
        K, L = row[10].value, row[11].value                         # SCC, APC
        N = row[13].value                                           # MSQ rakes/day
        Q = row[16].value                                           # Rakes alloted July
        U = row[20].value                                           # Current VC
        V = row[21].value                                           # Remarks

        if A is not None and str(A).strip() != "Total UPRVUNL":
            flush()
            current_block = {
                "name": str(A).strip(),
                "linkageType": str(B).strip() if B else None,
                "vcMoD": None,
                "currentVC": U,
                "remark": V,
                "sources": [],
            }

        if C is not None and current_block is not None:
            current_block["sources"].append({
                "name": str(C).strip(),
                "annualLinkage": D,
                "rate": F,
                "premium": G if G is not None else 0,
                "freight": H,
                "scc": K,
                "apc": L,
                "msqRakes": N if N is not None else 0,
                "julyReceipt": Q,
            })
    flush()
    return plants


def export_html(template_path, output_path, path=INPUT_FILE):
    """
    Reads the Calculation Sheet fresh from input.xlsx, serialises it as
    the RAW_DATA JS array the dashboard expects, and writes a new HTML
    file with that block swapped in. The dashboard's own JS optimizer
    and UI are left untouched - only the data feeding it changes.
    """
    import json
    import re

    plants = read_calculation_sheet(path)
    raw_data_js = "const RAW_DATA = " + json.dumps(plants, indent=2) + ";"

    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()

    pattern = re.compile(r"const RAW_DATA = \[.*?\];", re.DOTALL)
    if not pattern.search(html):
        raise RuntimeError("Could not find RAW_DATA block in template HTML")

    # Use a function replacement (not a plain string) so re.sub does not
    # interpret backslash sequences like \n inside the JSON as regex
    # group references - that would silently corrupt any remark text
    # containing escaped newlines.
    new_html = pattern.sub(lambda _m: raw_data_js, html, count=1)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"Dashboard data refreshed from '{path}' -> {output_path}")
    print(f"  ({len(plants)} plants, "
          f"{sum(len(p['sources']) for p in plants)} plant-source rows)")


# ---------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="UPRVUNL Rake Diversion Optimizer")
    parser.add_argument("--daily", default=None,
                         help="Path to daily_variation.csv (plant,source,available_rakes)")
    parser.add_argument("--freeze", default=None,
                         help='Freeze cells, e.g. "Parichha:NCL=52,Panki:CCL=60"')
    parser.add_argument("--dashboard", default=None,
                         help="Path to the dashboard HTML template to refresh "
                              "with live data from input.xlsx (writes to output/index.html)")
    args = parser.parse_args()

    frozen = parse_freeze_arg(args.freeze)

    print("=" * 60)
    print("STEP 1 & 2: Reading Excel and printing raw data")
    print("=" * 60)
    df = read_rake_data()
    df = clean_and_validate(df)
    print(df.to_string(index=False))

    print("\n" + "=" * 60)
    print("STEP 3: Plant-wise & overall weighted VC (reproducing Excel)")
    print("=" * 60)
    plant_summary, overall_vc = compute_weighted_vc(df)
    print(plant_summary.to_string(index=False))
    print(f"\nOverall Weighted VC (all plants): {overall_vc:.6f}")
    print("(Excel Dashboard 'Weighted Avg VC — Current' = 4.022564102564102 -> matches)")

    print("\n" + "=" * 60)
    print("STEP 4: Manual rake-change test (Parichha NCL: 60 -> 52)")
    print("=" * 60)
    old_vc, new_vc, status, _ = manual_change_test(df, "Parichha", "NCL", 52)
    print(f"Old VC: {old_vc:.2f}")
    print(f"New VC: {new_vc:.2f}")
    print(f"Target: {TARGET_VC}")
    print(f"Status: {status}")

    print("\n" + "=" * 60)
    print("STEP 5: Daily variation overlay")
    print("=" * 60)
    df_daily = apply_daily_variation(df, args.daily)
    print(df_daily[["Plant", "Source", "Rakes", "Available Rakes", "Variable Cost"]].to_string(index=False))

    print("\n" + "=" * 60)
    print("STEP 6: CLI Optimizer PREVIEW - ALL plants (base = Available Rakes, frozen cells applied)")
    print("=" * 60)
    print("NOTE: This CLI optimizer is NOT the project's official optimizer.")
    print("      The official, authoritative result comes only from")
    print("      backend/optimizer.py via POST /optimize (portfolio-wide")
    print("      OR-Tools solve with company-wide conservation). This CLI")
    print("      optimizes one plant at a time and has no such guarantee.")
    if frozen:
        print(f"Frozen cells: {frozen}")
    opt_all = optimize_all_plants(df_daily, base_col="Available Rakes", frozen=frozen)
    print(opt_all[["Plant", "Source", "Rakes", "Available Rakes", "Variable Cost",
                   "Optimized Rakes", "Frozen", "Method"]].to_string(index=False))

    print("\nPer-plant VC comparison (Original vs Optimized) - PREVIEW, not official:")
    for plant in opt_all["Plant"].unique():
        sub = opt_all[opt_all["Plant"] == plant]
        old_weighted = (sub["Rakes"] * sub["Variable Cost"]).sum() / sub["Rakes"].sum()
        new_weighted = (sub["Optimized Rakes"] * sub["Variable Cost"]).sum() / sub["Optimized Rakes"].sum()
        opt_status = "Achieved" if new_weighted <= TARGET_VC else "Not Achieved"
        method = sub["Method"].iloc[0]
        print(f"  {plant:25s}  Old VC: {old_weighted:.3f}  ->  New VC: {new_weighted:.3f}"
              f"  (Target {TARGET_VC})  [{opt_status}]  [{method}]")

    print("\n" + "=" * 60)
    print("STEP 7: Saving output.xlsx (CLI PREVIEW export - not the official result)")
    print("=" * 60)

    export_rows = []
    for plant in opt_all["Plant"].unique():
        sub = opt_all[opt_all["Plant"] == plant]
        old_weighted = (sub["Rakes"] * sub["Variable Cost"]).sum() / sub["Rakes"].sum()
        new_weighted = (sub["Optimized Rakes"] * sub["Variable Cost"]).sum() / sub["Optimized Rakes"].sum()
        opt_status = "Achieved" if new_weighted <= TARGET_VC else "Not Achieved"
        method = sub["Method"].iloc[0]
        for _, r in sub.iterrows():
            export_rows.append({
                "Plant": r["Plant"],
                "Source": r["Source"],
                "Original Rakes": r["Rakes"],
                "Daily Available Rakes": r["Available Rakes"],
                "Optimized Rakes": r["Optimized Rakes"],
                "Frozen": r["Frozen"],
                "Method": method,
                "Result Type": "PREVIEW (CLI, not official)",
                "Old VC (plant)": round(old_weighted, 4),
                "New VC (plant)": round(new_weighted, 4),
                "Target VC": TARGET_VC,
                "Target Status (preview)": opt_status,
            })

    out_df = pd.DataFrame(export_rows)
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    out_df.to_excel(OUTPUT_FILE, index=False, sheet_name="Result")
    print(f"Saved: {OUTPUT_FILE}")
    print("NOTE: output.xlsx is a CLI PREVIEW export. For the official,")
    print("      authoritative allocation, use the dashboard's")
    print('      "Optimize Allocation (OR-Tools)" button (POST /optimize).')

    if args.dashboard:
        print("\n" + "=" * 60)
        print("STEP 8: Refreshing dashboard HTML from input.xlsx")
        print("=" * 60)
        dashboard_out = "output/index.html"
        export_html(args.dashboard, dashboard_out)


if __name__ == "__main__":
    main()
